import csv
import openpyxl
from tqdm import tqdm
from openpyxl.styles import PatternFill

class DrugNamesChecker:
    def __init__(self, csv_file):
        self.russian_names = set()
        self.english_names = set()
        self.latin_names = set()
        self.load_from_csv(csv_file)

    def load_from_csv(self, csv_file):
        try:
            with open(csv_file, 'r', encoding='utf-8') as file:
                csv_reader = csv.reader(file, delimiter=';')
                for row in csv_reader:
                    if len(row) >= 3:
                        russian_name, english_name, latin_name = map(str.strip, row[:3])
                        self.russian_names.add(russian_name.lower())
                        self.english_names.add(english_name.lower())
                        self.latin_names.add(latin_name.lower())
                    else:
                        print(f"Skipping row: {row} - not enough columns")
        except FileNotFoundError:
            print(f"File '{csv_file}' not found.")
        except Exception as e:
            print(f"An error occurred while reading the CSV file: {str(e)}")

    def check_excel_file(self, excel_file, sheet_name="Общие формы", cmingrd_column_number=18):
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook[sheet_name]
            max_row = sheet.max_row

            for row in tqdm(sheet.iter_rows(min_row=3, min_col=cmingrd_column_number, max_col=cmingrd_column_number), total=max_row-2, desc="Processing Rows"):
                for cell in row:
                    if cell.value:  # Check if the cell has a value
                        value = str(cell.value).lower().strip()
                        if not any(value in names_set for names_set in [self.russian_names, self.english_names, self.latin_names]):
                            cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            
            workbook.save(excel_file)
        except FileNotFoundError:
            print(f"Excel file '{excel_file}' not found.")
        except KeyError:
            print(f"Sheet '{sheet_name}' not found in the Excel file.")
        except Exception as e:
            print(f"An error occurred while processing the Excel file: {str(e)}")
